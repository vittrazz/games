import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Шаг 1: Загружаем данные из листа "TDSheet"

file_path = 'Исходный файл.xlsx'
sheet_name = 'TDSheet'
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Шаг 2: Преобразовать таблицу в реляционную таблицу

# Получение заголовков из строк 2 и 3
headers = df.iloc[2].fillna('') + ' ' + df.iloc[3].fillna('')
headers = headers.str.strip().str.replace(' +', ' ', regex=True).str.replace('Unnamed: \d+_level_\d+', '', regex=True)

# Создание новой таблицы с заголовками
flat_df = pd.DataFrame(df.values[4:], columns=headers)
flat_df.dropna(how='all', inplace=True)

# Преобразование названий столбцов
flat_df.columns = pd.Series(flat_df.columns).str.strip().str.replace(' +', '_', regex=True)

# Печать названий столбцов для проверки
print("Колонки после преобразования:")
print(flat_df.columns)

# Проверка данных
print("Первые строки данных:")
print(flat_df.head())

# Шаг 3: Создать сводную таблицу
# Поиск столбцов с количеством документов
count_columns = [col for col in flat_df.columns if col.endswith('_Количество')]

# Создание сводной таблицы
pivot_table = flat_df.melt(id_vars=['МО_Статус'], value_vars=count_columns, var_name='Документ', value_name='Количество')
pivot_table['Статус'] = pivot_table['Документ'].apply(lambda x: 'Отправлено' if 'Отправлено' in x else 'Зарегистрировано')

# Печать промежуточных данных
print("Промежуточные данные после melt:")
print(pivot_table.head())

# Преобразование значений
pivot_table['Количество'] = pd.to_numeric(pivot_table['Количество'], errors='coerce').fillna(0) / 2

# Группировка данных для сводной таблицы
pivot_table = pivot_table.pivot_table(
    index='МО_Статус',
    columns='Статус',
    values='Количество',
    aggfunc='sum',
    fill_value=0
)

# Применение метода infer_objects для подавления предупреждения
pivot_table = pivot_table.infer_objects()

# Шаг 4: Добавить вычисляемое поле
pivot_table['Отношение_зарегистрированных_к_отправленным'] = (
    pivot_table.get('Зарегистрировано', 0) / pivot_table.get('Отправлено', 1)
).fillna(0) / 100
pivot_table['Отношение_зарегистрированных_к_отправленным'] = pivot_table['Отношение_зарегистрированных_к_отправленным'].apply(lambda x: "{:.2%}".format(x))

# Проверка данных сводной таблицы
print("Сводная таблица:")
print(pivot_table.head())

# Сохранение сводной таблицы в новый файл
new_file_path = 'Исходный файл обработанный.xlsx'
with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    flat_df.to_excel(writer, sheet_name='FlatTable', index=False)
    pivot_table.to_excel(writer, sheet_name='PivotTable')

# Шаг 5: Сохранение шагов в новый лист
steps = [
    "Загрузить данные из листа 'TDSheet'.",
    "Определить заголовки: Использовать строки 2 и 3 для создания заголовков.",
    "Создать новую таблицу: Преобразовать данные в реляционную таблицу, удалив пустые строки.",
    "Создать сводную таблицу на основе реляционной таблицы.",
    "Добавить вычисляемое поле в сводную таблицу и создать диаграмму."
]
steps_df = pd.DataFrame({
    "Шаг": range(1, len(steps) + 1),
    "Описание": steps
})

with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
    steps_df.to_excel(writer, sheet_name='Steps', index=False)

# Шаг 6: Создание диаграммы
wb = load_workbook(new_file_path)
ws = wb['PivotTable']

# Создание диаграммы
chart = BarChart()
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Количество зарегистрированных СЭМД"
chart.x_axis.title = "СЭМД"
chart.y_axis.title = "Количество"

# Добавление диаграммы на лист
ws.add_chart(chart, "E5")

# Сохранение файла с диаграммой
wb.save(new_file_path)

print("Процесс завершен.")
